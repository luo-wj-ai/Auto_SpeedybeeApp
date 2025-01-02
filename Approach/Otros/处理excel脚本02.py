from openpyxl import load_workbook

# 加载 Excel 文件
workbook = load_workbook(r"C:\Users\28449\Downloads\工具箱具体分类.xlsx")
sheet = workbook["统计"]  # 打开工作表“统计”

# 创建一个空字典用于分组
data_dict = {}

# 遍历表格，从第 2 行开始读取（忽略第一行）
for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
    product_name, feature = row  # 分别获取第一列和第二列的值
    if feature not in data_dict:
        data_dict[feature] = []  # 如果功能点不存在，初始化为空列表
    data_dict[feature].append(product_name)  # 添加产品名到对应功能点列表

# 打印结果
for feature, products in data_dict.items():
    unique_products = sorted(set(products))  # 去重并排序
    print(f"{feature}（{len(unique_products)}个型号）：{', '.join(unique_products)}")
